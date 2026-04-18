import json
import os
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import urlencode, urlparse

import schedule
from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


@dataclass
class Settings:
    excel_file_path: str
    excel_sheet_name: str
    panel_url: str
    history_url_base: str
    panel_username: str
    panel_password: str
    headless: bool
    max_links_per_batch: int
    retry_count: int
    retry_delay_seconds: float
    login_wait_seconds: int
    action_delay_seconds: float
    submission_wait_seconds: float
    post_submit_action: str
    submitted_sheet_name: str
    run_interval_minutes: int
    cookies_file: str


@dataclass
class FetchResult:
    links: list[str]
    valid_source_rows: list[int]


class AutomationError(RuntimeError):
    pass


def parse_bool(value: str, default: bool = False) -> bool:
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def load_settings() -> Settings:
    load_dotenv()

    settings = Settings(
        excel_file_path=os.getenv("EXCEL_FILE_PATH", "links.xlsx"),
        excel_sheet_name=os.getenv("EXCEL_SHEET_NAME", "Sheet1"),
        panel_url=os.getenv("PANEL_URL", "https://fast-index.icu/botfarms/panel.php"),
        history_url_base=os.getenv(
            "HISTORY_URL_BASE", "https://fast-index.icu/botfarms/enemy/check_all.php"
        ),
        panel_username=os.getenv("PANEL_USERNAME", "").strip(),
        panel_password=os.getenv("PANEL_PASSWORD", "").strip(),
        headless=parse_bool(os.getenv("HEADLESS", "true"), default=True),
        max_links_per_batch=int(os.getenv("MAX_LINKS_PER_BATCH", "500")),
        retry_count=int(os.getenv("RETRY_COUNT", "3")),
        retry_delay_seconds=float(os.getenv("RETRY_DELAY_SECONDS", "3")),
        login_wait_seconds=int(os.getenv("LOGIN_WAIT_SECONDS", "25")),
        action_delay_seconds=float(os.getenv("ACTION_DELAY_SECONDS", "1.5")),
        submission_wait_seconds=float(os.getenv("SUBMISSION_WAIT_SECONDS", "6")),
        post_submit_action=os.getenv("POST_SUBMIT_ACTION", "none").strip().lower(),
        submitted_sheet_name=os.getenv("SUBMITTED_SHEET_NAME", "Submitted"),
        run_interval_minutes=int(os.getenv("RUN_INTERVAL_MINUTES", "0")),
        cookies_file=os.getenv("COOKIES_FILE", "cookies.json"),
    )

    missing = []
    if not settings.excel_file_path:
        missing.append("EXCEL_FILE_PATH")
    if not settings.panel_username:
        missing.append("PANEL_USERNAME")
    if not settings.panel_password:
        missing.append("PANEL_PASSWORD")
    if missing:
        raise AutomationError(f"Missing required environment values: {', '.join(missing)}")

    if settings.post_submit_action not in {"none", "clear", "move"}:
        raise AutomationError("POST_SUBMIT_ACTION must be one of: none, clear, move")

    excel_path = Path(settings.excel_file_path)
    if not excel_path.exists():
        raise AutomationError(f"Excel file not found: {settings.excel_file_path}")

    return settings


def chunked(items: list[str], size: int) -> Iterable[list[str]]:
    for i in range(0, len(items), size):
        yield items[i : i + size]


def build_history_url(settings: Settings) -> str:
    query = urlencode(
        {
            "username": settings.panel_username,
            "password": settings.panel_password,
        }
    )
    return f"{settings.history_url_base}?{query}"


def is_valid_url(value: str) -> bool:
    try:
        parsed = urlparse(value)
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)
    except Exception:
        return False


def fetch_links_from_excel(settings: Settings) -> FetchResult:
    workbook = load_workbook(settings.excel_file_path)
    if settings.excel_sheet_name not in workbook.sheetnames:
        workbook.close()
        raise AutomationError(
            f"Worksheet '{settings.excel_sheet_name}' not found in {settings.excel_file_path}"
        )

    worksheet = workbook[settings.excel_sheet_name]

    seen = set()
    unique_links: list[str] = []
    valid_source_rows: list[int] = []

    for row_number in range(1, worksheet.max_row + 1):
        raw_value = worksheet.cell(row=row_number, column=1).value
        value = "" if raw_value is None else str(raw_value).strip()

        if not value:
            continue
        if not is_valid_url(value):
            continue
        if value in seen:
            continue

        seen.add(value)
        unique_links.append(value)
        valid_source_rows.append(row_number)

    workbook.close()

    print(f"Fetched {len(unique_links)} links from Excel")
    return FetchResult(links=unique_links, valid_source_rows=valid_source_rows)


def setup_driver(headless: bool) -> webdriver.Chrome:
    options = ChromeOptions()
    options.add_argument("--window-size=1600,1000")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    if headless:
        options.add_argument("--headless=new")

    try:
        service = Service(ChromeDriverManager().install())
        return webdriver.Chrome(service=service, options=options)
    except Exception:
        return webdriver.Chrome(options=options)


def retry(action: Callable[[], None], retries: int, delay_seconds: float, label: str) -> None:
    for attempt in range(1, retries + 1):
        try:
            action()
            return
        except Exception as exc:
            if attempt == retries:
                raise AutomationError(f"Failed: {label}. Last error: {exc}") from exc
            print(f"Retry {attempt}/{retries} for '{label}' after error: {exc}")
            time.sleep(delay_seconds)


def wait_and_find_any(driver: webdriver.Chrome, selectors: list[tuple[str, str]], timeout: int):
    end_time = time.time() + timeout
    last_error = None

    while time.time() < end_time:
        for by, locator in selectors:
            try:
                element = WebDriverWait(driver, 1).until(EC.presence_of_element_located((by, locator)))
                return element
            except Exception as err:
                last_error = err
        time.sleep(0.25)

    raise TimeoutException(f"Could not locate any expected element. Last error: {last_error}")


def cookies_path(settings: Settings) -> Path:
    return Path(settings.cookies_file)


def save_cookies(driver: webdriver.Chrome, settings: Settings) -> None:
    data = driver.get_cookies()
    path = cookies_path(settings)
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def load_cookies(driver: webdriver.Chrome, settings: Settings) -> bool:
    path = cookies_path(settings)
    if not path.exists():
        return False

    try:
        cookies = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return False

    driver.get(settings.panel_url)
    for cookie in cookies:
        c = dict(cookie)
        c.pop("sameSite", None)
        try:
            driver.add_cookie(c)
        except Exception:
            continue

    driver.refresh()
    return True


def is_logged_in(driver: webdriver.Chrome) -> bool:
    textarea_selectors = [
        (By.XPATH, "//textarea[contains(@placeholder, 'Paste Links')]"),
        (By.XPATH, "//label[contains(., 'Paste Links')]/following::textarea[1]"),
        (By.XPATH, "//textarea[contains(@name, 'link') or contains(@id, 'link')]"),
    ]
    try:
        wait_and_find_any(driver, textarea_selectors, timeout=4)
        return True
    except TimeoutException:
        return False


def login_if_needed(driver: webdriver.Chrome, settings: Settings) -> None:
    def _attempt_login() -> None:
        print("Opening panel...")
        driver.get(settings.panel_url)

        loaded = load_cookies(driver, settings)
        if loaded and is_logged_in(driver):
            print("Already logged in (cookies/session reused)")
            return

        print("Logging in...")

        user_input = wait_and_find_any(
            driver,
            selectors=[
                (By.CSS_SELECTOR, "input[name='username']"),
                (By.CSS_SELECTOR, "input[id='username']"),
                (By.CSS_SELECTOR, "input[type='text']"),
                (By.XPATH, "//input[contains(@placeholder, 'User') or contains(@placeholder, 'Email') ]"),
            ],
            timeout=settings.login_wait_seconds,
        )

        password_input = wait_and_find_any(
            driver,
            selectors=[
                (By.CSS_SELECTOR, "input[name='password']"),
                (By.CSS_SELECTOR, "input[id='password']"),
                (By.CSS_SELECTOR, "input[type='password']"),
            ],
            timeout=settings.login_wait_seconds,
        )

        user_input.clear()
        user_input.send_keys(settings.panel_username)
        password_input.clear()
        password_input.send_keys(settings.panel_password)

        submit_button = wait_and_find_any(
            driver,
            selectors=[
                (By.CSS_SELECTOR, "button[type='submit']"),
                (By.CSS_SELECTOR, "input[type='submit']"),
                (By.XPATH, "//button[contains(., 'Login') or contains(., 'Sign in') or contains(., 'Submit')]"),
            ],
            timeout=10,
        )
        submit_button.click()

        WebDriverWait(driver, settings.login_wait_seconds).until(lambda d: is_logged_in(d))

        time.sleep(settings.action_delay_seconds)
        save_cookies(driver, settings)
        print("Login successful")

    retry(_attempt_login, settings.retry_count, settings.retry_delay_seconds, "login")


def get_textarea(driver: webdriver.Chrome):
    return wait_and_find_any(
        driver,
        selectors=[
            (By.XPATH, "//textarea[contains(@placeholder, 'Paste Links')]"),
            (By.XPATH, "//label[contains(., 'Paste Links')]/following::textarea[1]"),
            (By.XPATH, "//textarea[contains(@name, 'link') or contains(@id, 'link')]"),
            (By.CSS_SELECTOR, "textarea"),
        ],
        timeout=20,
    )


def get_import_button(driver: webdriver.Chrome):
    return wait_and_find_any(
        driver,
        selectors=[
            (By.XPATH, "//button[contains(., 'Import Links')]"),
            (By.XPATH, "//input[@type='submit' and contains(@value, 'Import Links')]"),
            (By.XPATH, "//button[contains(., 'Import')]"),
        ],
        timeout=20,
    )


def submit_links_batch(
    driver: webdriver.Chrome,
    settings: Settings,
    links_batch: list[str],
    batch_no: int,
    total_batches: int,
) -> None:
    def _submit() -> None:
        print(f"Submitting batch {batch_no}/{total_batches} with {len(links_batch)} links...")

        textarea = get_textarea(driver)
        textarea.click()
        textarea.send_keys(Keys.CONTROL, "a")
        textarea.send_keys(Keys.DELETE)
        time.sleep(settings.action_delay_seconds)

        textarea.send_keys("\n".join(links_batch))
        time.sleep(settings.action_delay_seconds)

        import_button = get_import_button(driver)
        import_button.click()

        time.sleep(settings.submission_wait_seconds)
        print(f"Batch {batch_no} submitted successfully")

    retry(_submit, settings.retry_count, settings.retry_delay_seconds, f"submit batch {batch_no}")


def post_submit_update(settings: Settings, source_rows: list[int], submitted_links: list[str]) -> None:
    if settings.post_submit_action == "none" or not source_rows:
        return

    workbook = load_workbook(settings.excel_file_path)
    if settings.excel_sheet_name not in workbook.sheetnames:
        workbook.close()
        raise AutomationError(
            f"Worksheet '{settings.excel_sheet_name}' not found in {settings.excel_file_path}"
        )

    source_sheet = workbook[settings.excel_sheet_name]

    if settings.post_submit_action == "clear":
        for row in source_rows:
            source_sheet.cell(row=row, column=1, value=None)
        workbook.save(settings.excel_file_path)
        workbook.close()
        print(f"Cleared {len(source_rows)} submitted rows from Excel column A")
        return

    if settings.post_submit_action == "move":
        if settings.submitted_sheet_name in workbook.sheetnames:
            submitted_sheet = workbook[settings.submitted_sheet_name]
        else:
            submitted_sheet = workbook.create_sheet(settings.submitted_sheet_name)

        next_row = submitted_sheet.max_row + 1
        if next_row == 1 and submitted_sheet.cell(1, 1).value in (None, ""):
            next_row = 1

        for link in submitted_links:
            submitted_sheet.cell(row=next_row, column=1, value=link)
            next_row += 1

        for row in source_rows:
            source_sheet.cell(row=row, column=1, value=None)

        workbook.save(settings.excel_file_path)
        workbook.close()
        print(
            f"Moved {len(submitted_links)} links to sheet '{settings.submitted_sheet_name}' and cleared source rows"
        )


def run_once() -> None:
    settings = load_settings()
    print(f"History URL: {build_history_url(settings)}")

    fetched = fetch_links_from_excel(settings)

    if not fetched.links:
        print("No valid links found, nothing to submit")
        return

    all_batches = list(chunked(fetched.links, settings.max_links_per_batch))
    print(f"Prepared {len(all_batches)} batch(es)")

    driver = setup_driver(settings.headless)
    try:
        login_if_needed(driver, settings)
        time.sleep(settings.action_delay_seconds)

        for idx, batch in enumerate(all_batches, start=1):
            submit_links_batch(driver, settings, batch, idx, len(all_batches))

        post_submit_update(settings, fetched.valid_source_rows, fetched.links)
    finally:
        driver.quit()


def main() -> None:
    settings = load_settings()

    if settings.run_interval_minutes > 0:
        print(f"Scheduler enabled. Running every {settings.run_interval_minutes} minute(s)")
        run_once()
        schedule.every(settings.run_interval_minutes).minutes.do(run_once)
        while True:
            schedule.run_pending()
            time.sleep(1)
    else:
        run_once()


if __name__ == "__main__":
    main()
